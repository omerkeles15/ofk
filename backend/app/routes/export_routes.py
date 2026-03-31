from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import Optional
from datetime import datetime
import io, json

from app.database import get_db
from app.models import Device, DeviceData, IOPointHistory

router = APIRouter(prefix="/api", tags=["export"])


@router.get("/export/{device_id}")
async def export_device_data(
    device_id: str,
    from_ts: Optional[str] = Query(None, alias="from"),
    to_ts: Optional[str] = Query(None, alias="to"),
    db: AsyncSession = Depends(get_db),
):
    """Cihaz verilerini Excel olarak dışa aktar."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Cihaz bilgisi
    dev_r = await db.execute(select(Device).where(Device.id == device_id))
    device = dev_r.scalar_one_or_none()
    if not device:
        return {"error": "Cihaz bulunamadı"}

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    if device.device_type == "sensor":
        # Sensör — tek sayfa
        ws = wb.active
        ws.title = f"{device_id} Veriler"
        headers = ["#", "Değer", "Birim", "Durum", "Tarih/Saat"]
        style_header(ws, headers)

        where = [DeviceData.device_id == device_id]
        if from_ts:
            where.append(DeviceData.timestamp >= from_ts)
        if to_ts:
            where.append(DeviceData.timestamp <= to_ts)

        rows = (await db.execute(
            select(DeviceData).where(*where).order_by(DeviceData.timestamp.desc())
        )).scalars().all()

        for i, r in enumerate(rows):
            data = json.loads(r.data_json) if r.data_json else {}
            row_num = i + 2
            ws.cell(row=row_num, column=1, value=i + 1).border = thin_border
            ws.cell(row=row_num, column=2, value=data.get("value", "")).border = thin_border
            ws.cell(row=row_num, column=3, value=data.get("unit", "")).border = thin_border
            ws.cell(row=row_num, column=4, value=data.get("status", "")).border = thin_border
            ws.cell(row=row_num, column=5, value=r.timestamp or "").border = thin_border

        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 18

    else:
        # PLC — her I/O noktası ayrı sayfa
        ws = wb.active
        ws.title = "Özet"
        ws.cell(row=1, column=1, value=f"Cihaz: {device_id}").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Tag: {device.tag_name}")
        ws.cell(row=3, column=1, value=f"Tip: {device.device_type} / {device.subtype}")
        ws.cell(row=4, column=1, value=f"Dışa Aktarım: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

        # Tüm benzersiz adresleri bul
        addr_r = await db.execute(
            select(IOPointHistory.address).where(IOPointHistory.device_id == device_id).distinct()
        )
        addresses = sorted([r[0] for r in addr_r.all()])

        for addr in addresses:
            safe_name = addr.replace("/", "_")[:31]
            ws_io = wb.create_sheet(title=safe_name)
            headers = ["#", "Değer", "Tarih/Saat"]
            style_header(ws_io, headers)

            where = [IOPointHistory.device_id == device_id, IOPointHistory.address == addr]
            if from_ts:
                where.append(IOPointHistory.timestamp >= from_ts)
            if to_ts:
                where.append(IOPointHistory.timestamp <= to_ts)

            rows = (await db.execute(
                select(IOPointHistory).where(*where).order_by(IOPointHistory.timestamp.desc())
            )).scalars().all()

            for i, r in enumerate(rows):
                row_num = i + 2
                ws_io.cell(row=row_num, column=1, value=i + 1).border = thin_border
                ws_io.cell(row=row_num, column=2, value=r.value).border = thin_border
                ws_io.cell(row=row_num, column=3, value=r.timestamp or "").border = thin_border

            for col in range(1, 4):
                ws_io.column_dimensions[chr(64 + col)].width = 20

    # Excel dosyasını memory'de oluştur
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{device_id}_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
